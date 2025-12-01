# info_csv.py
import io
from openpyxl import Workbook

from info_table import display_name_from_mention


def build_logsummary_csv_bytes(donations, supplies, ledger, id_to_name=None):
    """
    Build an Excel (.xlsx) file that mirrors your 4 sections:
      - Donations Breakdown  (sheet: Donations)
      - Overall Totals       (sheet: Totals)
      - Supply Mission       (sheet: Supplies)
      - Ledger Transactions  (sheet: Ledger)

    Returns XLSX data as bytes, ready for discord.File().
    (Function name is kept for compatibility with existing imports.)
    """

    # --------- Aggregate donations like before ----------
    donation_map = {}
    for d in donations:
        nm = d["name"]
        mt = d["materials"]
        if nm not in donation_map:
            donation_map[nm] = {"count": 0, "total": 0.0}
        donation_map[nm]["count"] += 1
        donation_map[nm]["total"] += mt

    sorted_don = sorted(
        donation_map.items(),
        key=lambda kv: (-kv[1]["count"], -kv[1]["total"])
    )

    total_count = sum(v["count"] for v in donation_map.values())
    total_mat = sum(v["total"] for v in donation_map.values())

    # --------- Create workbook ----------
    wb = Workbook()

    # By default, Workbook() creates one sheet; we'll reuse it for Donations
    ws_don = wb.active
    ws_don.title = "Donations"

    # Donations sheet header
    ws_don.append(["Name", "Donations", "Total Materials Value"])
    for name, stats in sorted_don:
        disp_name = display_name_from_mention(name, id_to_name)
        ws_don.append([
            disp_name,
            stats["count"],
            round(stats["total"], 2),
        ])

    # Totals sheet
    ws_tot = wb.create_sheet(title="Totals")
    ws_tot.append(["Total Donations", "Total Materials Value"])
    ws_tot.append([total_count, round(total_mat, 2)])

    # Supplies sheet
    ws_sup = wb.create_sheet(title="Supplies")
    ws_sup.append(["Name", "Supplies Delivered"])
    for s in supplies:
        disp_name = display_name_from_mention(s["name"], id_to_name)
        ws_sup.append([disp_name, round(s["amount"], 2)])

    # Ledger sheet
    ws_led = wb.create_sheet(title="Ledger")
    ws_led.append(["Name", "Transition", "Amount"])
    for l in ledger:
        disp_name = display_name_from_mention(l["name"], id_to_name)
        transition = "Deposit" if l["transition"] == "Deposit" else "Withdrawal"
        ws_led.append([disp_name, transition, round(l["amount"], 2)])

    # --------- Save workbook to bytes ----------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
